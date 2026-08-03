"""
processor.py
Supplier Performance Report Tool
Part 1
---------------------------------
• Imports
• Validation
• Reading Excel files
• Cleaning data
• Merge Purchase Register & Receiving Report
• Delivery Days calculation
"""
from pathlib import Path
from io import BytesIO

import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)

from openpyxl.utils import get_column_letter

from openpyxl.chart import (
    BarChart,
    PieChart,
    Reference
)

from openpyxl.chart.label import DataLabelList

from openpyxl.formatting.rule import (
    ColorScaleRule
)

from config import (

    REPORT_COLUMNS,

    REGISTER_REQUIRED_COLUMNS,

    RECEIVING_INPUT_COLUMNS,

    PURCHASE_REGISTER_COLUMNS,

    PURCHASE_RECEIVING_COLUMNS,

    EXCLUDED_SUPPLIERS,

    REPORT_TITLE,

    MASTER_SHEET,

    OUTPUT_FILE,

    HEADER_FILL,

    HEADER_FONT,

    TOTAL_FILL,

    FREEZE_PANES,

    HEADER_ROW_HEIGHT,

    DEFAULT_ROW_HEIGHT

)

# =============================================================================
# READ EXCEL FILE
# =============================================================================

def read_excel_file(file_path: str | Path) -> pd.DataFrame:
    """
    Reads Materials Control Excel exports (.xls, .xlsx, .xlsm)
    from the 'Data' worksheet.
    """

    file_path = Path(file_path)
    suffix = file_path.suffix.lower()

    if suffix == ".xls":

        df = pd.read_excel(
            file_path,
            sheet_name="Data",
            engine="xlrd"
        )

    elif suffix in [".xlsx", ".xlsm"]:

        df = pd.read_excel(
            file_path,
            sheet_name="Data",
            engine="openpyxl"
        )

    else:

        raise ValueError(
            f"Unsupported file type: {suffix}"
        )

    # Clean column names
    df.columns = (
        df.columns
          .astype(str)
          .str.strip()
          .str.replace(r"\s+", " ", regex=True)
    )

    return df

# =============================================================================
# VALIDATION
# =============================================================================

def validate_columns(df: pd.DataFrame, required: set, file_name: str):
    """
    Validate uploaded workbook columns.
    """

    missing = required - set(df.columns)

    if missing:
        raise ValueError(
            f"\n{file_name}\n\n"
            f"Missing columns:\n"
            f"{', '.join(sorted(missing))}"
        )



# =============================================================================
# READ PURCHASE REGISTER
# =============================================================================

def read_purchase_register(file_path: str | Path) -> pd.DataFrame:
    """
    Read Purchase Register workbook.
    """

    df = read_excel_file(file_path)

    validate_columns(
        df,
        PURCHASE_REGISTER_COLUMNS,
        "Purchase Register"
    )

    df = df[REGISTER_REQUIRED_COLUMNS].copy()

    df["Order No."] = (
        df["Order No."]
        .fillna("")
        .astype(str)
        .str.upper()
        .str.strip()
    )

    df["Order Date"] = pd.to_datetime(
        df["Order Date"],
        errors="coerce"
    )

    return df


# =============================================================================
# READ RECEIVING REPORT
# =============================================================================

def read_receiving_report(file_path: str | Path) -> pd.DataFrame:
    """
    Read Purchase Receiving Deviation report.
    """

    df = read_excel_file(file_path)

    validate_columns(
        df,
        PURCHASE_RECEIVING_COLUMNS,
        "Purchase Receiving Deviation"
    )

    return df


# =============================================================================
# CLEAN RECEIVING DATA
# =============================================================================

def clean_receiving_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    Keep required columns and convert datatypes.
    """

    df = df[RECEIVING_INPUT_COLUMNS].copy()

    df["Supplier"] = (
        df["Supplier"]
        .fillna("")
        .astype(str)
        .str.upper()
        .str.strip()
    )

    df["Order No."] = (
        df["Order No."]
        .fillna("")
        .astype(str)
        .str.upper()
        .str.strip()
    )

    df["Delivery Date"] = pd.to_datetime(
        df["Delivery Date"],
        errors="coerce"
    )

    numeric_columns = [

        "Ordered",
        "Booked QTY",
        "Variance QTY",
        "PO Price",
        "Booked Price",
        "Variance Price",
        "Variance Value"

    ]

    for col in numeric_columns:

        df[col] = pd.to_numeric(
            df[col],
            errors="coerce"
        ).fillna(0)

    return df


# =============================================================================
# REMOVE EXCLUDED SUPPLIERS
# =============================================================================

def remove_excluded_suppliers(df: pd.DataFrame) -> pd.DataFrame:
    """
    Remove suppliers defined in config.py.
    """

    return df[
        ~df["Supplier"].isin(EXCLUDED_SUPPLIERS)
    ].copy()


# =============================================================================
# MERGE ORDER DATES
# =============================================================================

def merge_order_dates(
    receiving_df: pd.DataFrame,
    register_df: pd.DataFrame
) -> pd.DataFrame:
    """
    Merge Order Date from Purchase Register into the
    Purchase Receiving Deviation report.

    Prevents row multiplication by ensuring only one
    Order Date exists per Order No.

    Ignores placeholder Order Numbers such as
    'No PO defined', blank values, N/A, etc.
    """

    # ---------------------------------------------------------
    # Clean Order Numbers
    # ---------------------------------------------------------

    register_df = register_df.copy()

    register_df["Order No."] = (
        register_df["Order No."]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    # ---------------------------------------------------------
    # Ignore invalid Order Numbers
    # ---------------------------------------------------------

    invalid_orders = {
        "",
        "NO PO DEFINED",
        "N/A",
        "NONE"
    }

    valid_register = register_df[
        ~register_df["Order No."]
        .str.upper()
        .isin(invalid_orders)
    ].copy()

    # ---------------------------------------------------------
    # Check for conflicting Order Dates
    # ---------------------------------------------------------

    conflicting = (
        valid_register
        .groupby("Order No.")["Order Date"]
        .nunique()
    )

    conflicting = conflicting[conflicting > 1]

    if not conflicting.empty:

        conflicting_orders = ", ".join(conflicting.index.astype(str))

        raise ValueError(
            "Data quality issue detected.\n\n"
            "The following Order Numbers have multiple "
            f"Order Dates in the Purchase Register:\n\n"
            f"{conflicting_orders}\n\n"
            "Please verify the Purchase Register export."
        )

    # ---------------------------------------------------------
    # Keep one Order Date per Order Number
    # ---------------------------------------------------------

    order_lookup = (
        valid_register[["Order No.", "Order Date"]]
        .drop_duplicates(subset="Order No.")
    )

    # ---------------------------------------------------------
    # Merge
    # ---------------------------------------------------------

    merged = receiving_df.merge(
        order_lookup,
        how="left",
        on="Order No."
    )

    return merged

# =============================================================================
# CALCULATE DELIVERY DAYS
# =============================================================================

def calculate_delivery_days(df: pd.DataFrame) -> pd.DataFrame:
    """
    Delivery Days = Delivery Date - Order Date
    """

    df["Delivery Days"] = (
        df["Delivery Date"] -
        df["Order Date"]
    ).dt.days

    return df


# =============================================================================
# PREPARE REPORT DATASET
# =============================================================================

def prepare_report_data(
    purchase_register_file: str | Path,
    receiving_report_file: str | Path
) -> pd.DataFrame:
    """
    Complete preprocessing pipeline.
    """

    register = read_purchase_register(
        purchase_register_file
    )

    receiving = read_receiving_report(
        receiving_report_file
    )

    receiving = clean_receiving_data(receiving)

    receiving = remove_excluded_suppliers(receiving)

    merged = merge_order_dates(
        receiving,
        register
    )

    merged = calculate_delivery_days(merged)

    merged = merged[REPORT_COLUMNS]

    merged.sort_values(
        ["Supplier", "Order Date", "Order No."],
        inplace=True
    )

    merged.reset_index(
        drop=True,
        inplace=True
    )

    return merged  

# =============================================================================
# MASTER SUMMARY KPI CALCULATIONS
# =============================================================================

def create_master_summary(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create supplier performance summary.
    Average Delivery Days is calculated per unique
    Purchase Order instead of per article line.
    """

    # -------------------------------------------------------------------------
    # Delivery Days (One record per Purchase Order)
    # -------------------------------------------------------------------------

    delivery_summary = (
        df[
            ["Supplier", "Order No.", "Delivery Days"]
        ]
        .drop_duplicates(
            subset=["Supplier", "Order No."]
        )
        .groupby(
            "Supplier",
            as_index=False
        )
        .agg(
            Average_Delivery_Days=("Delivery Days", "mean")
        )
    )

    # -------------------------------------------------------------------------
    # Main Supplier KPIs
    # -------------------------------------------------------------------------

    summary = (
        df.groupby("Supplier", as_index=False)
        .agg(
            Orders=("Order No.", "nunique"),
            Ordered_Qty=("Ordered", "sum"),
            Received_Qty=("Booked QTY", "sum"),
            Qty_Variance=("Variance QTY", "sum"),
            Price_Variance=("Variance Value", "sum")
        )
    )

    # -------------------------------------------------------------------------
    # Merge Average Delivery Days
    # -------------------------------------------------------------------------

    summary = summary.merge(
        delivery_summary,
        on="Supplier",
        how="left"
    )

    # -------------------------------------------------------------------------
    # Order Fulfillment Rate
    # -------------------------------------------------------------------------

    summary["Order Fulfillment Rate %"] = (
        (
            summary["Received_Qty"] /
            summary["Ordered_Qty"]
        )
        .replace([float("inf")], 0)
        .fillna(0)
        .round(4)
    )

    # -------------------------------------------------------------------------
    # Rename Columns
    # -------------------------------------------------------------------------

    summary.rename(
        columns={
            "Ordered_Qty": "Ordered Qty",
            "Received_Qty": "Received Qty",
            "Qty_Variance": "Qty Variance",
            "Price_Variance": "Price Variance",
            "Average_Delivery_Days": "Average Delivery Days"
        },
        inplace=True
    )

    # -------------------------------------------------------------------------
    # Round Values
    # -------------------------------------------------------------------------

    summary["Average Delivery Days"] = (
        summary["Average Delivery Days"]
        .round(1)
    )

    # -------------------------------------------------------------------------
    # Sort Suppliers
    # -------------------------------------------------------------------------

    summary.sort_values(
        by=[
            "Order Fulfillment Rate %",
            "Average Delivery Days"
        ],
        ascending=[
            False,
            True
        ],
        inplace=True
    )

    summary.reset_index(
        drop=True,
        inplace=True
    )

    return summary
    
# =============================================================================
# EXECUTIVE SUMMARY
# =============================================================================

def create_executive_summary(df: pd.DataFrame) -> dict:
    """
    Dashboard KPI cards.
    """

    ordered = df["Ordered"].sum()
    received = df["Booked QTY"].sum()

    fill_rate = (
        (received / ordered) 
        if ordered else 0
    )

    return {

        "Total Suppliers":
            df["Supplier"].nunique(),

        "Total Orders":
            df["Order No."].nunique(),

        "Total Ordered Qty":
            ordered,

        "Total Received Qty":
            received,

        "Overall Order Fulfillment Rate %":
            round(fill_rate, 4),

        "Average Delivery Days":
            round(df["Delivery Days"].mean(), 1),

        "Total Price Variance":
            df["Variance Value"].sum(),

        "Total Quantity Variance":
            df["Variance QTY"].sum()

    }

# =============================================================================
# MASTER SUMMARY SHEET
# =============================================================================


from openpyxl.styles import Font

def write_master_summary(workbook, summary_df):

    ws = workbook.create_sheet("Master Summary")

    # -----------------------------------------------------
    # Headers
    # -----------------------------------------------------

    ws.append(summary_df.columns.tolist())

    # -----------------------------------------------------
    # Write Summary
    # -----------------------------------------------------

    current_row = 2

    for row in summary_df.itertuples(index=False):

        ws.append(list(row))

        supplier_cell = ws.cell(current_row, 1)

        supplier_name = str(supplier_cell.value)

        supplier_cell.hyperlink = (
            f"#'{supplier_name[:31]}'!A1"
        )

        supplier_cell.style = "Hyperlink"

        current_row += 1

    # -----------------------------------------------------
    # Header Dictionary
    # -----------------------------------------------------

    headers = {
        cell.value: cell.column
        for cell in ws[1]
    }

    # -----------------------------------------------------
    # Format Order Fulfillment Rate as %
    # -----------------------------------------------------

    if "Order Fulfillment Rate %" in headers:

        fulfillment_col = headers["Order Fulfillment Rate %"]

        for row in range(2, ws.max_row + 1):

            ws.cell(
                row,
                fulfillment_col
            ).number_format = "0.00%"

    return ws

# =============================================================================
# SUPPLIER KPI PANEL
# =============================================================================

def supplier_kpis(df: pd.DataFrame):

    ordered = df["Ordered"].sum()

    received = df["Booked QTY"].sum()

    fill_rate = (
        (received / ordered)
        if ordered else 0
    )

    delivery_days = (
        df[
            ["Order No.", "Delivery Days"]
        ]
        .drop_duplicates(subset=["Order No."])
        ["Delivery Days"]
        .mean()
    )

    return [

        ("Orders", df["Order No."].nunique()),

        ("Ordered Qty", ordered),

        ("Received Qty", received),

        ("Order Fulfillment Rate %", round(fill_rate, 4)),

        ("Quantity Variance", df["Variance QTY"].sum()),

        ("Price Variance", df["Variance Value"].sum()),

        (
            "Average Delivery Days",
            round(delivery_days, 1)
        )

    ]


###############################################################################
# ARTICLE SUMMARY (PIVOT STYLE)
###############################################################################

def create_article_summary(sheet, supplier_df, start_row):
    """
    Creates a pivot-style Monthly Article Summary with
    expandable Order Number details.

    Returns
    -------
    summary_row : int
        First row of the article summary table.

    summary_rows : list[int]
        Worksheet row numbers containing ONLY the
        article summary rows. Used for chart creation.
    """

    # -------------------------------------------------------------------------
    # Title
    # -------------------------------------------------------------------------

    sheet.cell(start_row, 1).value = "Monthly Article Summary"

    start_row += 2

    # -------------------------------------------------------------------------
    # Headers
    # -------------------------------------------------------------------------

    headers = [

        "Article",
        "Ordered Qty",
        "Delivered Qty",
        "Qty Variance",
        "No. of Orders"

    ]

    for col, header in enumerate(headers, start=1):

        sheet.cell(start_row, col).value = header

    summary_row = start_row + 1

    # -------------------------------------------------------------------------
    # Keep track of ONLY article rows
    # -------------------------------------------------------------------------

    summary_rows = []

    # -------------------------------------------------------------------------
    # Monthly Article Totals
    # -------------------------------------------------------------------------

    article_summary = (

        supplier_df

        .groupby("Article", as_index=False)

        .agg(

            Ordered=("Ordered", "sum"),

            Delivered=("Booked QTY", "sum"),

            Variance=("Variance QTY", "sum"),
            
            Order_Frequency=("Order No.", "nunique")

        )

        .sort_values("Article")

    )

    current_row = summary_row

    # -------------------------------------------------------------------------
    # Write each Article
    # -------------------------------------------------------------------------

    for _, article in article_summary.iterrows():

        article_row = current_row

        # -------------------------------------------------------------
        # Summary Row
        # -------------------------------------------------------------

        sheet.cell(current_row, 1).value = article["Article"]
        sheet.cell(current_row, 2).value = article["Ordered"]
        sheet.cell(current_row, 3).value = article["Delivered"]
        sheet.cell(current_row, 4).value = article["Variance"]

        freq_cell = sheet.cell(current_row, 5)
        freq_cell.value = article["Order_Frequency"]
        freq_cell.number_format = "0"
        
        # Save this row for charting
        summary_rows.append(current_row)

        current_row += 1

        # -------------------------------------------------------------
        # Order Detail Rows
        # -------------------------------------------------------------

        detail = (

            supplier_df[
                supplier_df["Article"] == article["Article"]
            ]

            .sort_values(
                [
                    "Order Date",
                    "Order No."
                ]
            )

        )

        for _, order in detail.iterrows():

            sheet.cell(
                current_row,
                1
            ).value = "    " + str(order["Order No."])

            sheet.cell(
                current_row,
                2
            ).value = order["Ordered"]

            sheet.cell(
                current_row,
                3
            ).value = order["Booked QTY"]

            sheet.cell(
                current_row,
                4
            ).value = order["Variance QTY"]

            # Make order rows collapsible
            sheet.row_dimensions[current_row].outlineLevel = 1
            sheet.row_dimensions[current_row].hidden = True

            current_row += 1

        # Collapse under article
        sheet.row_dimensions[article_row].collapsed = True

    # -------------------------------------------------------------------------
    # Return ONLY article rows
    # -------------------------------------------------------------------------

    return summary_row, summary_rows

###############################################################################
# HELPER TABLE
###############################################################################

def create_helper_table(sheet, supplier_df, start_row):
    """
    Creates a hidden helper table containing one row per Purchase Order.
    Used for Excel KPI formulas.
    """

    helper = (
        supplier_df
        .groupby("Order No.", as_index=False )
        .agg(
            Order_Date=("Order Date", "first"),
            Last_Delivery_Date=("Delivery Date", "max")
        )
    )

    helper["Delivery Days"] = (
        helper["Last_Delivery_Date"]
        - helper["Order_Date"]
    ).dt.days

    sheet.cell(start_row, 27).value = "Order No."
    sheet.cell(start_row, 28).value = "Delivery Days"

    row = start_row + 1

    for _, order in helper.iterrows():

        sheet.cell(row, 27).value = order["Order No."]
        sheet.cell(row, 28).value = order["Delivery Days"]

        row += 1

    # Hide helper columns (AA = 27, AB = 28)
    sheet.column_dimensions["AA"].hidden = True
    sheet.column_dimensions["AB"].hidden = True

    return row - 1


# =============================================================================
# SUPPLIER WORKSHEETS
# =============================================================================

def create_supplier_sheets(workbook, report_df, worksheet_last_rows):

    suppliers = sorted(
        report_df["Supplier"].unique()
    )

    for supplier in suppliers:

        sheet = workbook.create_sheet(
            supplier[:31]
        )

        supplier_df = (
            report_df[
                report_df["Supplier"] == supplier
            ]
            .sort_values(
                [
                    "Order Date",
                    "Order No."
                ]
            )
        )

        # -----------------------------
        # Transaction Table
        # -----------------------------

        sheet.append(
            supplier_df.columns.tolist()
        )

        for row in supplier_df.itertuples(index=False):

            sheet.append(list(row))

        last_data_row = sheet.max_row

        worksheet_last_rows[sheet.title] = last_data_row
        
        # -----------------------------
        # HELPER TABLE
        # -----------------------------

        helper_start = last_data_row + 3

        helper_end = create_helper_table(
            sheet,
            supplier_df,
            helper_start
        )

        # -----------------------------
        # KPI PANEL
        # -----------------------------

        start_row = sheet.max_row + 3

        sheet.cell(
            start_row,
            1,
            "Supplier KPI Summary"
        )

        kpis = supplier_kpis( 
            supplier_df
        )

        for i, (kpi, value) in enumerate(
            kpis,
            start=start_row + 1
        ):

            sheet.cell(i, 1).value = kpi

            value_cell = sheet.cell(i, 2)

            if kpi == "Orders":

                value_cell.value = (
                    f"=COUNTA(AA{helper_start+1}:AA{helper_end})"
                )

            elif kpi == "Ordered Qty":

                value_cell.value = (
                    f"=SUM(H2:H{last_data_row})"
                )

            elif kpi == "Received Qty":

                value_cell.value = (
                    f"=SUM(J2:J{last_data_row})"
                )

            elif kpi == "Quantity Variance":

                value_cell.value = (
                    f"=SUM(K2:K{last_data_row})"
                )

            elif kpi == "Price Variance":

                value_cell.value = (
                    f"=SUM(N2:N{last_data_row})"
                )

            elif kpi == "Order Fulfillment Rate %":

                value_cell.value = (
                    f"=IF(B{start_row+2}=0,0,B{start_row+3}/B{start_row+2})"
                )

                value_cell.number_format = "0.00%"

            elif kpi == "Average Delivery Days":

                value_cell.value = (
                    f"=AVERAGE(AB{helper_start+1}:AB{helper_end})"
                )

                value_cell.number_format = "0.0"
        
        summary_start = start_row + len(kpis) + 4

        article_start, summary_rows = create_article_summary(
            sheet,
            supplier_df,
            summary_start
        )

        add_supplier_chart(
            sheet,
            article_start,
            summary_rows
        )
       

# =============================================================================
# BUILD WORKBOOK
# =============================================================================

def build_workbook(report_df):

    wb = Workbook()

    wb.remove(wb.active)

    summary = create_master_summary(
        report_df
    )

    write_master_summary(
        wb,
        summary
    )

    worksheet_last_rows = {}
    
    create_supplier_sheets(
        wb,
        report_df,
        worksheet_last_rows,
    )

    return wb 

###############################################################################
# EXCEL FORMATTING
###############################################################################

def format_worksheet(ws, last_data_row):
    """
    Apply professional worksheet formatting.

    Only the transaction table (rows 1 to last_data_row)
    receives column-specific formatting.
    """

    # -------------------------------------------------------------------------
    # Styles
    # -------------------------------------------------------------------------

    header_fill = PatternFill(
        fill_type="solid",
        fgColor=HEADER_FILL
    )

    header_font = Font(
        bold=True,
        color=HEADER_FONT
    )

    thin = Side(style="thin")

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    # -------------------------------------------------------------------------
    # Worksheet settings
    # -------------------------------------------------------------------------

    ws.freeze_panes = FREEZE_PANES
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{last_data_row}"

    # -------------------------------------------------------------------------
    # Transaction Table Header
    # -------------------------------------------------------------------------

    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT

    for cell in ws[1]:

        cell.fill = header_fill
        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        cell.border = border

    # -------------------------------------------------------------------------
    # Store transaction headers
    # -------------------------------------------------------------------------

    headers = {
        cell.column: str(cell.value)
        for cell in ws[1]
    }

    # -------------------------------------------------------------------------
    # Format ONLY the transaction table
    # -------------------------------------------------------------------------

    for row in ws.iter_rows(
        min_row=2,
        max_row=last_data_row
    ):

        ws.row_dimensions[row[0].row].height = DEFAULT_ROW_HEIGHT

        for cell in row:

            cell.border = border

            header = headers.get(cell.column, "")

            # Dates
            if "Date" in header:

                cell.number_format = "dd-mmm-yyyy"

            # Percentages
            elif "%" in header:

                cell.number_format = "0.00%"

            # Numbers
            elif isinstance(cell.value, (int, float)):

                cell.number_format = "#,##0.00"

    # -------------------------------------------------------------------------
    # Auto-fit all columns
    # -------------------------------------------------------------------------

    for column in ws.columns:

        max_length = 0

        column_letter = get_column_letter(column[0].column)

        for cell in column:

            try:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            except Exception:

                pass

        ws.column_dimensions[column_letter].width = min(
            max_length + 3,
            40
        )

###############################################################################
# CONDITIONAL FORMATTING
###############################################################################

def apply_conditional_formatting(ws):

    headers = {

        cell.value: cell.column

        for cell in ws[1]

    }

    if "Delivery Days" in headers:

        col = get_column_letter(

            headers["Delivery Days"]

        )

        ws.conditional_formatting.add(

            f"{col}2:{col}{ws.max_row}",

            ColorScaleRule(

                start_type="min",

                start_color="63BE7B",

                mid_type="percentile",

                mid_value=50,

                mid_color="FFEB84",

                end_type="max",

                end_color="F8696B"

            )

        )


###############################################################################
# CHARTS
###############################################################################

def add_supplier_chart(
    ws,
    article_summary
):
    """
    Creates an Ordered vs Delivered chart using the
    article_summary DataFrame.

    A visible Chart Summary table is written first,
    followed immediately by the chart.
    """

    # ---------------------------------------------------------
    # CHART SUMMARY TITLE
    # ---------------------------------------------------------

    chart_title_row = ws.max_row + 3

    ws.cell(
        chart_title_row,
        1
    ).value = "Ordered vs Delivered Summary"

    # ---------------------------------------------------------
    # HEADERS
    # ---------------------------------------------------------

    header_row = chart_title_row + 1

    headers = [
        "Article",
        "Ordered Qty",
        "Delivered Qty"
    ]

    for col, header in enumerate(headers, start=1):

        cell = ws.cell(
            header_row,
            col
        )

        cell.value = header

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=HEADER_FILL
        )

        cell.font = Font(
            bold=True,
            color=HEADER_FONT
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    # ---------------------------------------------------------
    # WRITE SUMMARY DATA
    # ---------------------------------------------------------

    data_start = header_row + 1

    current_row = data_start

    for _, row in article_summary.iterrows():

        ws.cell(
            current_row,
            1
        ).value = row["Article"]

        ws.cell(
            current_row,
            2
        ).value = row["Ordered"]

        ws.cell(
            current_row,
            3
        ).value = row["Delivered"]

        current_row += 1

    # ---------------------------------------------------------
    # TOTAL ROW
    # ---------------------------------------------------------

    ws.cell(current_row, 1).value = "TOTAL"

    ws.cell(
        current_row,
        2
    ).value = (
        f"=SUM(B{data_start}:B{current_row-1})"
    )

    ws.cell(
        current_row,
        3
    ).value = (
        f"=SUM(C{data_start}:C{current_row-1})"
    )

    total_row = current_row

    # ---------------------------------------------------------
    # CREATE CHART
    # ---------------------------------------------------------

    chart = BarChart()

    chart.type = "col"

    chart.style = 10

    chart.title = "Ordered vs Delivered by Article"

    chart.y_axis.title = "Quantity"

    chart.x_axis.title = "Article"

    chart.height = 8

    chart.width = 16

    chart.dLbls = DataLabelList()

    chart.dLbls.showVal = True

    data = Reference(
        ws,
        min_col=2,
        max_col=3,
        min_row=header_row,
        max_row=total_row - 1
    )

    categories = Reference(
        ws,
        min_col=1,
        min_row=data_start,
        max_row=total_row - 1
    )

    chart.add_data(
        data,
        titles_from_data=True
    )

    chart.set_categories(
        categories
    )

    # ---------------------------------------------------------
    # POSITION CHART
    # ---------------------------------------------------------

    chart_row = total_row + 2

    ws.add_chart(
        chart,
        f"A{chart_row}"
    )

###############################################################################
# MASTER DASHBOARD
###############################################################################

from openpyxl.styles import Font

def add_dashboard(master_ws, report_df):
    """
    Creates a live Executive Dashboard using Excel formulas
    linked to the Master Summary table.
    """

    # ---------------------------------------------------------
    # Dashboard Title
    # ---------------------------------------------------------

    master_ws.insert_rows(1, amount=10)

    master_ws["A1"] = REPORT_TITLE

    master_ws["A1"].font = Font(
        bold=True,
        size=16
    )

    # ---------------------------------------------------------
    # Locate Master Summary columns
    # ---------------------------------------------------------

    headers = {
        cell.value: cell.column
        for cell in master_ws[11]
    }

    summary_last_row = master_ws.max_row

    # Convert column numbers to Excel letters

    from openpyxl.utils import get_column_letter

    supplier_col = get_column_letter(headers["Supplier"])
    orders_col = get_column_letter(headers["Orders"])
    ordered_col = get_column_letter(headers["Ordered Qty"])
    received_col = get_column_letter(headers["Received Qty"])
    qty_var_col = get_column_letter(headers["Qty Variance"])
    price_var_col = get_column_letter(headers["Price Variance"])
    avg_days_col = get_column_letter(headers["Average Delivery Days"])

    # ---------------------------------------------------------
    # Dashboard Labels
    # ---------------------------------------------------------

    dashboard = [

        ("Total Suppliers",
         f"=COUNTA({supplier_col}12:{supplier_col}{summary_last_row})"),

        ("Total Orders",
         f"=SUM({orders_col}12:{orders_col}{summary_last_row})"),

        ("Total Ordered Qty",
         f"=SUM({ordered_col}12:{ordered_col}{summary_last_row})"),

        ("Total Received Qty",
         f"=SUM({received_col}12:{received_col}{summary_last_row})"),

        ("Overall Order Fulfillment Rate",
         f"=IF(SUM({ordered_col}12:{ordered_col}{summary_last_row})=0,"
         f"0,"
         f"SUM({received_col}12:{received_col}{summary_last_row})/"
         f"SUM({ordered_col}12:{ordered_col}{summary_last_row}))"),

        ("Average Delivery Days",
         f"=AVERAGE({avg_days_col}12:{avg_days_col}{summary_last_row})"),

        ("Total Price Variance",
         f"=SUM({price_var_col}12:{price_var_col}{summary_last_row})"),

        ("Total Quantity Variance",
         f"=SUM({qty_var_col}12:{qty_var_col}{summary_last_row})")

    ]

    # ---------------------------------------------------------
    # Write Dashboard
    # ---------------------------------------------------------

    start_row = 2

    for label, formula in dashboard:

        master_ws.cell(
            start_row,
            1
        ).value = label

        value_cell = master_ws.cell(
            start_row,
            2
        )

        value_cell.value = formula

        # Percentage formatting
        if "Fulfillment Rate" in label:

            value_cell.number_format = "0.00%"

        # Decimal formatting
        elif "Average Delivery Days" in label:

            value_cell.number_format = "0.0"

        # Quantity formatting
        else:

            value_cell.number_format = '#,##0.00'

        start_row += 1

###############################################################################
# SAVE REPORT
###############################################################################
def save_workbook(workbook):
    """
    Save workbook to memory instead of disk.
    """

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return output


###############################################################################
# MAIN PROCESS
###############################################################################

def process_files(

    purchase_register,

    receiving_report,

):

    # ---------------------------------------------------------
    # Prepare report data
    # ---------------------------------------------------------

    report_df = prepare_report_data(

        purchase_register,

        receiving_report

    )

    # ---------------------------------------------------------
    # Determine report period for output filename
    # ---------------------------------------------------------

    report_period = (

        report_df["Delivery Date"]

        .dropna()

        .max()

        .strftime("%B_%Y")

    )

    # ---------------------------------------------------------
    # Build workbook
    # ---------------------------------------------------------

    workbook = build_workbook(

        report_df

    )

    # ---------------------------------------------------------
    # Add dashboard
    # ---------------------------------------------------------

    master = workbook[MASTER_SHEET]

    add_dashboard(

        master,

        report_df

    )

    # ---------------------------------------------------------
    # Format worksheets
    # ---------------------------------------------------------

    for sheet in workbook.worksheets:

        if sheet.title in worksheet_last_rows:

            format_worksheet(

                sheet,

                worksheet_last_rows[sheet.title]

            )

        apply_conditional_formatting(sheet)
    # ---------------------------------------------------------
    # Return workbook and report period
    # ---------------------------------------------------------

    return save_workbook(workbook), report_period
